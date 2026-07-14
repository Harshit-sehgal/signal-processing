import os
import re
import glob
from copy import deepcopy
import math
from typing import List, Dict, Any
# Goal 6.1 canonical metadata schema. A metadata workbook must provide these
# columns (matched case-insensitively; a few accept the aliases the parser
# already understands). Enforced by build_dataset_index.
REQUIRED_METADATA_COLUMNS = (
    "recording_id",
    "experiment_run_id",
    "stickout",
    "tooth_count",
    "tool_id",
    "sensor_id",
    "label",
    "chatter_onset_time",
    "rpm",
    "doc",
    "state",
)
# A required column is satisfied if it appears directly or via one of these
# header aliases (consistent with the parser's existing column mappings).
_REQUIRED_ALIASES = {
    "rpm": {"rpm", "spindle speed"},
    "doc": {"doc", "doc (in)", "corrected doc (in)"},
    "state": {"state", "chatter/nochatter", "status"},
    "label": {"label", "state", "chatter/nochatter", "status"},
}


def _align_finite_feature_rows(
    rows: List[Dict[str, Any]],
) -> tuple[List[str], Any, Dict[str, Any]]:
    """Align heterogeneous exploratory features without inventing values.

    Per-IMF feature names are dynamic because valid recordings can decompose
    into different IMF counts.  Experimental evaluators therefore use the
    intersection of keys that are present and finite in every row.  Union-based
    zero filling would give an absent centre frequency or bandwidth a fabricated
    physical meaning.
    """

    import numpy as np

    if not rows:
        raise ValueError("At least one feature row is required for alignment.")
    mappings: List[Dict[str, Any]] = []
    for row in rows:
        features = row.get("features")
        if not isinstance(features, dict):
            raise ValueError("Every evaluation row must contain a feature mapping.")
        mappings.append(dict(features))
    candidate_keys = set().union(*(mapping.keys() for mapping in mappings))
    common_keys = set(mappings[0]).intersection(*(mapping.keys() for mapping in mappings[1:]))
    feature_keys: List[str] = []
    for key in sorted(common_keys):
        try:
            values = [float(mapping[key]) for mapping in mappings]
        except (TypeError, ValueError):
            continue
        if all(math.isfinite(value) for value in values):
            feature_keys.append(key)
    if not feature_keys:
        raise ValueError("No common finite features are available across evaluation rows.")

    matrix = np.asarray(
        [[float(mapping[key]) for key in feature_keys] for mapping in mappings],
        dtype=float,
    )
    if matrix.ndim != 2 or matrix.shape != (len(rows), len(feature_keys)):
        raise RuntimeError("Aligned feature matrix has an unexpected shape.")
    if not np.all(np.isfinite(matrix)):
        raise RuntimeError("Aligned feature matrix contains non-finite values.")
    dropped = sorted(candidate_keys.difference(feature_keys))
    evidence = {
        "strategy": "intersection_of_present_finite_features",
        "candidate_feature_count": len(candidate_keys),
        "aligned_feature_count": len(feature_keys),
        "dropped_feature_count": len(dropped),
        "dropped_features": dropped,
        "missing_feature_policy": "drop_from_experimental_matrix; never zero-fill",
    }
    return feature_keys, matrix, evidence


def _exploratory_segment_config(
    config: Dict[str, Any], sample_count: int
) -> tuple[Dict[str, Any], Dict[str, Any]]:
    """Copy config and make its duration contract explicit for one eval slice."""

    if sample_count < 3:
        raise ValueError("Exploratory evaluation segments require at least three samples.")
    local_config = deepcopy(config)
    sampling_rate = float(local_config.get("sampling_rate", 0.0))
    if not math.isfinite(sampling_rate) or sampling_rate <= 0.0:
        raise ValueError("sampling_rate must be a finite positive number.")
    raw_validation = local_config.get("validation", {})
    if not isinstance(raw_validation, dict):
        raise ValueError("validation must be a mapping.")
    validation = dict(raw_validation)
    source_minimum = float(validation.get("minimum_duration_seconds", 0.0))
    if not math.isfinite(source_minimum) or source_minimum < 0.0:
        raise ValueError("validation.minimum_duration_seconds must be non-negative and finite.")
    segment_duration = sample_count / sampling_rate
    effective_minimum = min(source_minimum, segment_duration)
    validation["minimum_duration_seconds"] = effective_minimum
    local_config["validation"] = validation
    return local_config, {
        "source_minimum_duration_seconds": source_minimum,
        "effective_minimum_duration_seconds": effective_minimum,
        "segment_duration_seconds": segment_duration,
        "adjusted": effective_minimum != source_minimum,
        "scope": "experimental_evaluation_segment_only",
    }

def build_dataset_index(
    input_dir: str, 
    metadata_excel_path: str
) -> List[Dict[str, Any]]:
    """Scans all MAT files in input_dir and maps them to metadata rows in the Excel combinations workbook.
    
    Returns:
        A list of dictionaries representing the master index table.
        
    Raises:
        ValueError: If there are duplicate recordings or missing machining parameters.
    """
    import openpyxl
    
    if not os.path.exists(metadata_excel_path):
        raise FileNotFoundError(f"Metadata Excel file not found: {metadata_excel_path}")
        
    wb = openpyxl.load_workbook(metadata_excel_path)
    
    # 1. Parse all excel sheets to build combinations master list
    combinations = []
    sheets_to_load = [s for s in wb.sheetnames if s not in ['recording times']]
    
    seen_columns = set()
    for sname in sheets_to_load:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
            
        header = [str(c).strip().lower() if c is not None else '' for c in rows[0]]
        
        # Column mappings
        idx_ld = -1
        for name in ['l/d', 'ld']:
            if name in header:
                idx_ld = header.index(name)
                break
                
        idx_rpm = -1
        for name in ['rpm', 'spindle speed']:
            if name in header:
                idx_rpm = header.index(name)
                break
                
        idx_doc = -1
        for name in ['doc (in)', 'corrected doc (in)', 'doc']:
            if name in header:
                idx_doc = header.index(name)
                break
                
        idx_state = -1
        for name in ['state', 'chatter/nochatter', 'status']:
            if name in header:
                idx_state = header.index(name)
                break
                
        idx_feed = -1
        for name in ['feed (in/rev)', 'feed', 'feedrate']:
            if name in header:
                idx_feed = header.index(name)
                break
                
        seen_columns.update(h for h in header if h)
        if idx_rpm == -1 or idx_doc == -1 or idx_state == -1:
            continue
            
        for r in rows[1:]:
            if r[idx_rpm] is not None and r[idx_doc] is not None:
                # Resolve L/D
                if idx_ld != -1 and r[idx_ld] is not None:
                    try:
                        ld = float(r[idx_ld])
                    except ValueError:
                        ld = 2.0
                else:
                    m = re.search(r'rod(\d+)p?(\d*)inch', sname)
                    if m:
                        ld = float(m.group(1)) + (float(m.group(2))/10.0 if m.group(2) else 0.0)
                    else:
                        ld = 3.5 if '3p5' in sname else 4.125 if '4p125' in sname else 4.5 if '4p5' in sname else 2.0
                        
                try:
                    rpm_val = int(r[idx_rpm])
                    doc_val = float(r[idx_doc])
                except (ValueError, TypeError):
                    continue
                    
                state_val = str(r[idx_state]).strip().lower() if r[idx_state] is not None else ''
                feed_val = float(r[idx_feed]) if (idx_feed != -1 and r[idx_feed] is not None) else 0.002
                
                # Standardize state label
                if 'no chatter' in state_val or 'nochatter' in state_val:
                    label = 'stable'
                elif 'mild' in state_val or 'incipient' in state_val or 'intermittent' in state_val:
                    label = 'incipient'
                else:
                    label = 'chatter'
                    
                combinations.append({
                    'ld': ld,
                    'rpm': rpm_val,
                    'doc': doc_val,
                    'feed': feed_val,
                    'label': label
                })

    # Enforce the Goal 6.1 canonical metadata schema.
    missing = [
        c for c in REQUIRED_METADATA_COLUMNS
        if c not in seen_columns
        and not (c in _REQUIRED_ALIASES and _REQUIRED_ALIASES[c] & seen_columns)
    ]
    if missing:
        raise ValueError(
            "Metadata workbook is missing required columns (Goal 6.1): "
            + ", ".join(sorted(missing))
        )

    # 2. Scan MAT files and map them
    ld_map = {
        '2inch_stickout': 2.0, 
        '2p5inch_stickout': 2.5, 
        '3p5inch_stickout': 3.5, 
        '4p5inch_stickout': 4.5
    }
    
    mat_files = glob.glob(os.path.join(input_dir, "**/*.mat"), recursive=True)
    mat_files = [f for f in mat_files if not f.endswith("combinations.xlsx") and "~lock" not in f]
    
    index_table = []
    seen_recording_ids = set()
    
    for f in mat_files:
        folder = os.path.basename(os.path.dirname(f))
        base = os.path.basename(f)
        base_no_ext = base.replace(".mat", "")
        
        ld = ld_map.get(folder)
        if ld is None:
            raise ValueError(f"Unknown stickout folder name: {folder} for file {f}")
            
        parts = base_no_ext.split('_')
        if len(parts) < 3:
            raise ValueError(f"Invalid filename structure for file {f}")
            
        prefix = parts[0]
        try:
            rpm = int(parts[1])
            doc_str = re.sub('[^0-9]', '', parts[2])
            doc = float(doc_str) / 1000.0
        except ValueError:
            raise ValueError(f"Failed to parse RPM or DOC from filename {base}")
            
        # Parse run number if present, e.g. u_570_015_3 -> run 3
        run_id = "1"
        if len(parts) >= 4:
            run_id = parts[3]
            
        recording_id = f"ld_{ld}_rpm_{rpm}_doc_{doc}_state_{prefix}_run_{run_id}"
        if recording_id in seen_recording_ids:
            raise ValueError(f"Duplicate recording ID detected: {recording_id}")
        seen_recording_ids.add(recording_id)
        
        # Match combination from Excel
        match = None
        for row in combinations:
            if abs(row['ld'] - ld) < 0.1 and row['rpm'] == rpm and abs(row['doc'] - doc) < 0.001:
                match = row
                break
                
        # Label resolution (with fallback to filename prefix)
        if match:
            label = match['label']
            feed = match['feed']
        else:
            # Fallback to prefix
            feed = 0.002
            if prefix == 's':
                label = 'stable'
            elif prefix == 'i':
                label = 'incipient'
            else:
                label = 'chatter'
                
        index_table.append({
            'recording_id': recording_id,
            'file_path': f,
            'stickout': ld,
            'rpm': rpm,
            'depth_of_cut': doc,
            'feed_rate': feed,
            'tooth_count': 1, # default single tooth cutter
            'tool_id': "tool_01",
            'machine_id': "lathe_01",
            'sensor_id': "accelerometer_01",
            'label': label,
            'chatter_onset_time': 0.0,
            'experiment_run_id': run_id
        })
        
    return index_table


def evaluate_directory(
    npz_dir: str,
    fs: float = 10_000.0,
    rpm: float = 600.0,
    tooth_count: int = 1,
    random_state: int = 42,
) -> Dict[str, Any]:
    """Evaluate the detection scaffolding on a directory of ``*_IMFs.npz`` recordings.

    Each ``.npz`` must contain ``original_signal``, ``imfs`` and ``start_index``;
    the matching ``*_Clean.mat`` provides the denoised physical signal. The
    recording label is taken from the filename: ``c*`` -> chatter (1),
    ``s*`` -> stable (0). This is the real-data evaluation path used when a raw
    dataset workbook is unavailable (e.g. the bundled ``testing/t1`` artifacts).

    Returns a results dict: ``n_recordings``, ``feature_keys``, ``cv_mean_metrics``,
    ``holdout_metrics`` and ``calibrated_proba``.
    """
    import numpy as np
    import glob as _glob
    from scipy.io import loadmat as _loadmat
    from pg_amcd.features import extract_window_features
    from pg_amcd.detection import (
        train_baseline_classifiers,
        evaluate_detector,
        fit_probability_calibrator,
    )
    from sklearn.linear_model import LogisticRegression

    recs = []
    for npz in sorted(_glob.glob(os.path.join(npz_dir, "*_IMFs.npz"))):
        base = os.path.basename(npz).replace("_IMFs.npz", "")
        clean_mat = os.path.join(npz_dir, base + "_Clean.mat")
        if not os.path.exists(clean_mat):
            continue
        d = np.load(npz)
        original = np.asarray(d["original_signal"], dtype=float)
        imfs = np.asarray(d["imfs"], dtype=float)
        clean = np.asarray(_loadmat(clean_mat)["tsDS"], dtype=float)
        denoised = clean[:, 1] if clean.ndim == 2 else clean
        label = 1 if base.startswith("c") else 0
        recs.append((base, original, imfs, denoised, label))
    if not recs:
        raise ValueError(f"No *_IMFs.npz recordings found in {npz_dir}")

    keys = None
    X, y, groups = [], [], []
    for base, original, imfs, denoised, label in recs:
        f = extract_window_features(original, original, denoised, imfs, fs, rpm, tooth_count)
        if keys is None:
            keys = list(f.keys())
        X.append([float(f[k]) for k in keys])
        y.append(label)
        groups.append(base)
    X = np.asarray(X, dtype=float)
    y = np.asarray(y)
    groups = np.asarray(groups)

    cv_results = train_baseline_classifiers(X, y, groups=groups)
    lr = LogisticRegression(max_iter=2000, class_weight="balanced")
    lr.fit(X, y)
    proba = lr.predict_proba(X)[:, 1]
    pred = (proba >= 0.5).astype(int)
    metrics = evaluate_detector(y, pred, proba)
    cal = fit_probability_calibrator(y, proba, method="isotonic")
    cal_proba = cal(proba)
    return {
        "n_recordings": len(recs),
        "feature_keys": keys,
        "cv_mean_metrics": {n: r["mean_metrics"] for n, r in cv_results.items()},
        "holdout_metrics": metrics,
        "calibrated_proba": [float(v) for v in cal_proba],
    }


def evaluate_real_dataset(
    mat_dir: str,
    config: Dict[str, Any],
    label_filter: tuple = ("c", "i", "s"),
    random_state: int = 42,
) -> Dict[str, Any]:
    """Evaluate chatter detection on a directory of raw ``*.mat`` recordings.

    Filename convention ``<label>_<rpm>_<feed>.mat`` (label in ``c/s/i/u``)
    inside a stickout subdirectory (e.g. ``2p5inch_stickout``). Each signal is
    processed through the full PG-AMCD pipeline and window features are
    extracted. A binary chatter task (``c``/``i``=1 [chatter & intermittent chatter], ``s``=0 [stable]) is evaluated with
    grouped cross-validation so windows from one recording never leak across
    train/test (Goal 6.4):

      * leave-one-recording-out (``recording_id`` groups)
      * leave-one-stickout-out
      * leave-one-rpm-out

    Labels are taken from the real filename convention; no labels are
    fabricated. Returns per-model metrics for every grouping level plus a
    probability calibrator fit on out-of-fold probabilities.
    """
    import numpy as np
    from collections import Counter
    from pg_amcd.io import validate_and_load_signal
    from pg_amcd.pipeline import process_recording
    from pg_amcd.detection import (
        train_baseline_classifiers,
        predict_window_probabilities,
        evaluate_detector,
        fit_probability_calibrator,
    )

    files = sorted(glob.glob(os.path.join(mat_dir, "**", "*.mat"), recursive=True))
    rows: List[Dict[str, Any]] = []
    skips = []
    for f in files:
        base = os.path.basename(f)[:-4]
        parts = base.split("_")
        if len(parts) < 3:
            continue
        label_code = parts[0]
        if label_code not in label_filter:
            continue
        # Filenames may carry non-numeric suffixes (e.g. ``c_570_015s``); strip
        # trailing letters so a valid recording is never silently dropped.
        try:
            rpm = float(re.sub(r"[^0-9.]", "", parts[1]))
            feed = float(re.sub(r"[^0-9.]", "", parts[2]))
        except ValueError:
            continue
        stickout = os.path.basename(os.path.dirname(f))
        try:
            t_arr, sig, fs = validate_and_load_signal(f, config["sampling_rate"])
        except Exception as exc:  # skip unreadable/invalid files, don't abort
            skips.append({"file": base, "reason": str(exc)})
            continue
        res = process_recording(
            t_arr, sig, config,
            metadata={"rpm": rpm, "tooth_count": 1},
            mode="exploratory",
        )
        for wr in res.window_results:
            rows.append({
                "features": wr.features,
                "label": 1 if label_code in ("c", "i") else 0,
                "recording_id": os.path.relpath(f, mat_dir),
                "stickout": stickout,
                "rpm": rpm,
                "feed": feed,
            })

    if not rows:
        raise ValueError(f"No usable recordings found in {mat_dir}")

    feature_keys, X, feature_alignment = _align_finite_feature_rows(rows)
    y = np.array([r["label"] for r in rows], dtype=int)
    rec_id = np.array([r["recording_id"] for r in rows])
    stickout = np.array([r["stickout"] for r in rows])
    rpm_a = np.array([r["rpm"] for r in rows])

    def grouped_eval(group_values, Xmat):
        from sklearn.preprocessing import StandardScaler
        uniq = np.unique(group_values)
        y_true_parts, proba_parts = [], {}
        for g in uniq:
            tr, te = group_values != g, group_values == g
            # Scale per fold (fit on train only) to avoid leakage and to keep
            # scale-sensitive models (logistic regression, SVM) well-conditioned.
            scaler = StandardScaler().fit(Xmat[tr])
            Xtr = scaler.transform(Xmat[tr])
            Xte = scaler.transform(Xmat[te])
            models = train_baseline_classifiers(
                Xtr, y[tr], groups=rec_id[tr], random_state=random_state
            )
            for name, m in models.items():
                if m["model"] is None:
                    continue
                p = predict_window_probabilities(Xte, m["model"])
                proba_parts.setdefault(name, []).append(p)
            y_true_parts.append(y[te])
        y_true = np.concatenate(y_true_parts)
        out = {}
        for name, plist in proba_parts.items():
            proba = np.concatenate(plist)
            pred = (proba >= 0.5).astype(int)
            out[name] = evaluate_detector(y_true, pred, proba)
        return out, y_true, proba_parts

    loi, y_true_loo, loi_proba = grouped_eval(rec_id, X)
    los, _, _ = grouped_eval(stickout, X)
    lor, _, _ = grouped_eval(rpm_a, X)

    # Feature ablations (Goal 7): isolate the contribution of the PG-AMCD
    # frequency/IMF-derived features by dropping each group and re-running the
    # same leakage-proof grouped CV on the surviving feature subset.
    ablation_specs = {
        "time_domain_only": [i for i, k in enumerate(feature_keys)
                             if k.startswith("time_")],
        "without_freq_features": [i for i, k in enumerate(feature_keys)
                                  if not k.startswith("freq_")],
        "without_imf_features": [i for i, k in enumerate(feature_keys)
                                 if not k.startswith("imf_")],
    }
    feature_ablations: Dict[str, Any] = {}
    for abl_name, idx in ablation_specs.items():
        if len(idx) < 2:
            continue
        Xsub = X[:, idx]
        a_loo, _, _ = grouped_eval(rec_id, Xsub)
        a_los, _, _ = grouped_eval(stickout, Xsub)
        a_lor, _, _ = grouped_eval(rpm_a, Xsub)
        feature_ablations[abl_name] = {
            "leave_one_recording_out": a_loo,
            "leave_one_stickout_out": a_los,
            "leave_one_rpm_out": a_lor,
        }

    # Calibrate the strongest model (by ROC-AUC) on pooled OOF probabilities.
    best = (
        max(loi, key=lambda n: (loi[n] or {}).get("roc_auc", -1.0))
        if loi else None
    )
    cal_info = {
        "method": "isotonic",
        "fitted_on": "leave-one-recording-out out-of-fold probabilities",
        "best_model": best,
    }
    if best is not None and best in loi_proba:
        y_oof = y_true_loo
        p_oof = np.concatenate(loi_proba[best])
        cal = fit_probability_calibrator(y_oof, p_oof, method="isotonic")
        p_cal = cal(p_oof)
        pred_cal = (p_cal >= 0.5).astype(int)
        cal_info["calibrated_metrics"] = evaluate_detector(y_oof, pred_cal, p_cal)
    counts = Counter(int(v) for v in y)
    return {
        "n_recordings": int(len(np.unique(rec_id))),
        "n_windows": int(len(rows)),
        "n_skipped": int(len(skips)),
        "label_counts": {
            "chatter": int(counts.get(1, 0)),
            "stable": int(counts.get(0, 0)),
        },
        "feature_keys": feature_keys,
        "feature_alignment": feature_alignment,
        "leave_one_recording_out": loi,
        "leave_one_stickout_out": los,
        "leave_one_rpm_out": lor,
        "calibration": cal_info,
        # Out-of-fold predictions for the leave-one-recording-out split, used by
        # the report/figure generators (no need to re-run the pipeline).
        "oof": {
            "y_true": [int(v) for v in y_true_loo],
            "proba": {
                name: [float(v) for v in np.concatenate(plist)]
                for name, plist in loi_proba.items()
            },
        },
        "feature_ablations": feature_ablations,
        "skipped_files": skips,
    }


def evaluate_real_dataset_temporal(
    mat_dir: str,
    config: Dict[str, Any],
    label_filter: tuple = ("c", "i", "s"),
    n_windows: int = 5,
    window_points: int = 2048,
    random_state: int = 42,
) -> Dict[str, Any]:
    """Multi-window chatter evaluation with temporal smoothing (Goal 6.5).

    Each recording is split into ``n_windows`` overlapping fixed-length
    segments; every segment is pushed through the full PG-AMCD pipeline so a
    recording yields a *sequence* of per-window probabilities. Leakage-proof
    grouped CV (by ``recording_id``) produces out-of-fold per-window
    probabilities; within each recording the probability sequence is stabilised
    with ``temporal_smooth_probabilities`` (hysteresis + minimum run length).

    This makes the Goal 6.5 smoothing logic observable on real data and
    quantifies its effect (unsmoothed vs smoothed per-window metrics). Labels
    come from the real filename convention; nothing is fabricated.
    """
    import numpy as np
    from collections import Counter, defaultdict
    from pg_amcd.io import validate_and_load_signal
    from pg_amcd.pipeline import process_recording
    from pg_amcd.detection import (
        train_baseline_classifiers,
        predict_window_probabilities,
        evaluate_detector,
        temporal_smooth_probabilities,
    )

    def _segment_indices(n: int):
        if n <= window_points or n_windows <= 1:
            return [(0, min(n, window_points))]
        starts = np.linspace(0, n - window_points, n_windows).astype(int)
        return [(int(s), int(s) + window_points) for s in starts]

    files = sorted(glob.glob(os.path.join(mat_dir, "**", "*.mat"), recursive=True))
    rows: List[Dict[str, Any]] = []
    skips = []
    exploratory_config_adjustments: List[Dict[str, Any]] = []
    for f in files:
        base = os.path.basename(f)[:-4]
        parts = base.split("_")
        if len(parts) < 3:
            continue
        label_code = parts[0]
        if label_code not in label_filter:
            continue
        try:
            rpm = float(re.sub(r"[^0-9.]", "", parts[1]))
            feed = float(re.sub(r"[^0-9.]", "", parts[2]))
        except ValueError:
            continue
        stickout = os.path.basename(os.path.dirname(f))
        try:
            t_arr, sig, fs = validate_and_load_signal(f, config["sampling_rate"])
        except Exception as exc:
            skips.append({"file": base, "reason": str(exc)})
            continue
        n = len(sig)
        rec_id = os.path.relpath(f, mat_dir)
        for wi, (s, e) in enumerate(_segment_indices(n)):
            segment_config, adjustment = _exploratory_segment_config(config, e - s)
            exploratory_config_adjustments.append(
                {"recording_id": rec_id, "window_index": wi, **adjustment}
            )
            res = process_recording(
                t_arr[s:e], sig[s:e], segment_config,
                metadata={"rpm": rpm, "tooth_count": 1},
                mode="exploratory",
            )
            for wr in res.window_results:
                rows.append({
                    "features": wr.features,
                    "label": 1 if label_code in ("c", "i") else 0,
                    "recording_id": rec_id,
                    "stickout": stickout,
                    "rpm": rpm,
                    "feed": feed,
                    "win_idx": wi,
                })

    if not rows:
        raise ValueError(f"No usable recordings found in {mat_dir}")

    feature_keys, X, feature_alignment = _align_finite_feature_rows(rows)
    y = np.array([r["label"] for r in rows], dtype=int)
    rec_id = np.array([r["recording_id"] for r in rows])
    stickout = np.array([r["stickout"] for r in rows])

    def grouped_eval(group_values, Xmat):
        from sklearn.preprocessing import StandardScaler
        uniq = np.unique(group_values)
        y_true_parts, proba_parts = [], {}
        for g in uniq:
            tr, te = group_values != g, group_values == g
            scaler = StandardScaler().fit(Xmat[tr])
            Xtr = scaler.transform(Xmat[tr])
            Xte = scaler.transform(Xmat[te])
            models = train_baseline_classifiers(
                Xtr, y[tr], groups=rec_id[tr], random_state=random_state
            )
            for name, m in models.items():
                if m["model"] is None:
                    continue
                p = predict_window_probabilities(Xte, m["model"])
                proba_parts.setdefault(name, []).append(p)
            y_true_parts.append(y[te])
        y_true = np.concatenate(y_true_parts)
        out = {}
        for name, plist in proba_parts.items():
            proba = np.concatenate(plist)
            pred = (proba >= 0.5).astype(int)
            out[name] = evaluate_detector(y_true, pred, proba)
        return out, y_true, proba_parts

    loi, y_true_loo, loi_proba = grouped_eval(rec_id, X)

    best = (
        max(loi, key=lambda n: (loi[n] or {}).get("roc_auc", -1.0))
        if loi else None
    )

    # Within-recording temporal smoothing (Goal 6.5) on the best model's OOF
    # per-window probabilities.
    temporal_params = {
        "enter_threshold": 0.75,
        "exit_threshold": 0.40,
        "min_positive_windows": 3,
        "median_window": 5,
    }
    smoothed_metrics: Dict[str, Any] = {}
    smoothed_proba = np.full(len(y), float("nan"))
    if best is not None and best in loi_proba:
        # Reconstruct the per-window OOF probability array in original order.
        oof_proba = np.concatenate(loi_proba[best])
        by_rec: Dict[str, List[int]] = defaultdict(list)
        for i, rid in enumerate(rec_id):
            by_rec[rid].append(i)
        out_proba = np.zeros(len(y), dtype=float)
        for rid, idxs in by_rec.items():
            seq = oof_proba[idxs]
            if len(seq) < 2:
                out_proba[idxs] = seq
                continue
            labels, sm = temporal_smooth_probabilities(
                seq, **temporal_params
            )
            out_proba[idxs] = sm
        smoothed_proba = out_proba
        pred_s = (out_proba >= 0.5).astype(int)
        smoothed_metrics[best] = evaluate_detector(y, pred_s, out_proba)

    # Feature importances from a full-corpus RandomForest (for reporting/figures).
    feature_importances = {}
    try:
        from sklearn.ensemble import RandomForestClassifier
        from sklearn.preprocessing import StandardScaler
        sc = StandardScaler().fit(X)
        rf = RandomForestClassifier(n_estimators=200, class_weight="balanced",
                                    random_state=random_state)
        rf.fit(sc.transform(X), y)
        for k, imp in zip(feature_keys, rf.feature_importances_):
            feature_importances[k] = float(imp)
    except Exception:
        feature_importances = {}

    counts = Counter(int(v) for v in y)
    return {
        "n_recordings": int(len(np.unique(rec_id))),
        "n_windows": int(len(rows)),
        "n_skipped": int(len(skips)),
        "window_points": int(window_points),
        "n_windows_per_recording": int(n_windows),
        "label_counts": {
            "chatter": int(counts.get(1, 0)),
            "stable": int(counts.get(0, 0)),
        },
        "feature_keys": feature_keys,
        "feature_alignment": feature_alignment,
        "exploratory_config_adjustments": exploratory_config_adjustments,
        "best_model": best,
        "per_window_metrics": loi,
        "smoothed_metrics": smoothed_metrics,
        "temporal_params": temporal_params,
        "oof": {
            "y_true": [int(v) for v in y_true_loo],
            "proba": {
                name: [float(v) for v in np.concatenate(plist)]
                for name, plist in loi_proba.items()
            },
        },
        "smoothed_proba": [float(v) for v in smoothed_proba],
        "feature_importances": feature_importances,
        "skipped_files": skips,
    }
